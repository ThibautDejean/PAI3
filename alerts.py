import openpyxl as xl
import time
import winsound


def iter_retard_avion(): #changements de couleur des cases a  chaque boucle 
    wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
    feuille = wb['Vols en cours']
    colonne_heure_arrivée = []
    couleur_heure_arrivée = []

    #on récupère le temps actuel
    temps_actuel = time.strftime("%H:%M", time.localtime())
    print(temps_actuel)
    print('juste heure' , int(temps_actuel[0:2]))

    #on récupère les heures d'arrivée et la couleur de la case
    for col in feuille.iter_cols():
        header_cell = col[4]
        if header_cell.value == "Heure d'arrivée":
            for cell in col:
                if cell.value != None and cell.value != "Heure d'arrivée":
                    colonne_heure_arrivée.append(cell.value)
                    couleur_heure_arrivée.append(cell.fill.start_color.index[2:])

    #on compare les heures d'arrivée avec le temps actuel
    for heure in colonne_heure_arrivée:
        if (int(heure[0:2]) < int(temps_actuel[0:2])) or (int(heure[0:2]) == int(temps_actuel[0:2]) and int(heure[3:5]) < int(temps_actuel[3:5])):
            heure_retard = (int(temps_actuel[0:2]) - int(heure[0:2]))*60 + (int(temps_actuel[3:5]) - int(heure[3:5])) #en minutes
    
            if heure_retard > 15 : 
                #on regarde si la colonne n'est pas déjà coloriée en orange
                if couleur_heure_arrivée[colonne_heure_arrivée.index(heure)] != 'FFA500' and couleur_heure_arrivée[colonne_heure_arrivée.index(heure)] != 'FF0000' : 
                    #on colore la ligne en orange
                    fill = xl.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                    ligne = colonne_heure_arrivée.index(heure) + 1 + 5 #on ajoute 5 car les 5 premières lignes ne sont pas des vols et 1 car on commence à 0 en python
                    for j in range(1, feuille.max_column+1):
                        cell = feuille.cell(row=ligne, column=j)
                        cell.fill = fill
                    #on crée un son d'alarme 
                    winsound.PlaySound('alarm.wav', winsound.SND_FILENAME)
            if heure_retard > 30 : 
                #on regarde si pas déjà coloriée en rouge
                if couleur_heure_arrivée[colonne_heure_arrivée.index(heure)] != 'FF0000' : 
                    #on colore la ligne en rouge
                    fill = xl.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    ligne = colonne_heure_arrivée.index(heure) + 1 + 5
                    for j in range(1, feuille.max_column+1):
                        cell = feuille.cell(row=ligne, column=j)
                        cell.fill = fill
                    #on crée un son d'alarme 
                    winsound.PlaySound('alarm.wav', winsound.SND_FILENAME)
            else : 
                #on colore la ligne en vert
                fill = xl.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                ligne = colonne_heure_arrivée.index(heure) + 1 + 5
                for j in range(1, feuille.max_column+1):
                    cell = feuille.cell(row=ligne, column=j)
                    cell.fill = fill
        else : 
            #on colore la ligne en vert
            fill = xl.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            ligne = colonne_heure_arrivée.index(heure) + 1 + 5
            for j in range(1, feuille.max_column+1):
                cell = feuille.cell(row=ligne, column=j)
                cell.fill = fill

    wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')

