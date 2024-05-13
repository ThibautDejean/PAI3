import sqlite3
import openpyxl as xl

def plan_de_vol(corps,id_aeronef):                 
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()

        # Identifiant aeronef
        cur.execute('''REPLACE INTO "Plans de vols"(Aeronef) VALUES (?)''',(id_aeronef,))
        
        # Identifiant aerodrome de depart
        ligne=corps[4].split('-')
        depart=ligne[1]
        cur.execute('''UPDATE "Plans de vols" SET "Aerodrome de depart" = ? WHERE Aeronef = ?''',[(depart[0:5]),id_aeronef])

        # Heure de depart
        A=depart[5:10]
        B = A[0:3] + ':' + A[3:5]

        cur.execute('''UPDATE "Plans de vols" SET "Heure de depart" = ? WHERE Aeronef = ?''',[(B),id_aeronef])

        # Identifiant aerodrome d'arrivee
        ligne2=corps[8].split('-')
        arrivee=ligne2[1]
        cur.execute('''UPDATE "Plans de vols" SET "Aerodrome d'arrivee" = ? WHERE Aeronef = ?''',[(arrivee[0:5]),id_aeronef])

        # Duree du vol
        C=arrivee[5:10]
        D = C[0:3] + ':' + C[3:5]
        cur.execute('''UPDATE "Plans de vols" SET "Duree du vol" = ? WHERE Aeronef = ?''',[(D),id_aeronef])

        # Heure d'arrivee
        heure=int(depart[6:8])+int(arrivee[6:8])
        minute=int(depart[8:10])+int(arrivee[8:10])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)

        E = heure_arrivee[0:2] + ':' + heure_arrivee[2:4]
        
        cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ?''',[(E),id_aeronef])

        # Chemin
        
        ligne3 = corps[6].split(' ')
        ligne4 = ligne3[2:len(ligne3)]
        villes = ' '.join(ligne4)

        cur.execute('''UPDATE "Plans de vols" SET "Chemin" = ? WHERE Aeronef = ?''',[(villes),id_aeronef])

        print("declaration de plan de vol")
        conn.commit()
        conn.close()


def message_arrive(corps,id_aeronef):              
        
    # Identification de l'aeronef
    ligne=corps[0].split('-')
    b = ligne[1].split(' ')
    id = ' '+b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]+' '
    idbdd = b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]
    

    # Supression BDD

    conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
    cur = conn.cursor()

    cur.execute('''DELETE FROM "Plans de vols" WHERE Aeronef = ? ''', (idbdd,))

    conn.commit()
    conn.close()

    # Suppression ligne Excel

    wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
    feuille = wb['Vols en cours']

    for row in feuille.iter_rows():
            for cell in row :
                if str(cell.value) == str(id) : 
                    ligne = cell.row
                
    for j in range(4,11):
        feuille.cell(ligne,j).value = None
        feuille.cell(ligne,j).fill = xl.styles.PatternFill(fill_type=None)


    wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')