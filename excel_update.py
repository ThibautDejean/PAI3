import openpyxl as xl
import sqlite3

def ecriture_excel(corps, id_aeronef):              
        ### Fonction qui inscrit le mail dans le fichier Excel ###
        #Ouverture du fichier
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']
        #Ligne excel
        i=6
        while feuille.cell(i, 4).value != None :
            i+=1

        # Identifiant aerodrome de depart
        ligne=corps[4].split('-')
        depart=ligne[1]

        #Recuperation vol dans bdd
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''SELECT "Heure de depart","Duree du vol", "Aerodrome d'arrivee", "Heure d'arrivee", "Chemin" FROM "Plans de vols" WHERE Aeronef = ? ''', (id_aeronef,))
        
        vol=[]
        for ligne in cur.fetchall():
            vol=list(ligne)
        
        cur.close()
        conn.close()

        #Ecriture dans le fichier excel
        feuille.cell(i,4).value = id_aeronef
        feuille.cell(i,5).value = depart[0:5]
        feuille.cell(i,6).value = vol[0]
        feuille.cell(i,7).value = vol[1]
        feuille.cell(i,8).value = vol[2]
        feuille.cell(i,9).value = vol[3]
        feuille.cell(i,10).value = vol[4]

        #Sauvegarder
        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')

def message_delai(corps,id_aeronef):               

        #Base de donnee
        ligne=corps[4].split('-')
        depart=ligne[1]
        
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''UPDATE "Plans de vols" SET "Heure de depart" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))
        
        ligne2=corps[8].split('-')
        arrivee=ligne2[1]

        heure=int(depart[6:8])+int(arrivee[6:8])
        minute=int(depart[8:10])+int(arrivee[8:10])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)
        
        cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))

        conn.commit()
        conn.close()

        #Excel   
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']

        
        for row in feuille.iter_rows():
             for cell in row:
                 if cell.value == id_aeronef:
                     a = (cell.row,cell.column)
        
        feuille.cell(row=a[0],column=6).value=depart[5:10]

        feuille.cell(row=a[0],column=9).value=heure_arrivee

        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')


