import openpyxl as xl
import sqlite3

def message_changement(self,corps,id_aeronef):

    conn = sqlite3.connect('/Users/thibautdejean/Desktop/vols_pai.db')
    cur = conn.cursor()

    cur.execute('''SELECT "Duree du vol" FROM "Plans de vols WHERE Aeronef = ? ''', (id_aeronef,))
    duree=cur.fetchall()

    ligne=corps[0].split('-')
    depart = ligne[2]
    arrivee = ligne[3]
    

    heure=int(depart[6:8])+int(duree[0:2])
    minute=int(depart[8:10])+int(duree[2:4])

    if int(minute)>60:
        minute=int(minute)-60
        heure+=1
    heure_arrivee = str(heure)+str(minute)
    
    
    cur.execute('''UPDATE "Plans de vols" SET "Aerodrome de depart" = ? WHERE Aeronef = ? ''', (depart[5:10],id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Heure de départ" = ? WHERE Aeronef = ? ''', (depart[0:5],id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Aerodrome d'arrivee" = ? WHERE Aeronef = ? ''', (arrivee,id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ? ''', (heure_arrivee,id_aeronef))
    conn.commit()
    conn.close()

    #Excel   
    wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
    feuille = wb['Vols en cours']


    for row in feuille.iter_rows():
            for cell in row:
                if cell.value == id_aeronef:
                    a = (cell.row,cell.column)

    feuille.cell(row=a[0],column=5).value=depart[0:5]
    feuille.cell(row=a[0],column=6).value=depart[5:10]

    feuille.cell(row=a[0],column=8).value=arrivee
    feuille.cell(row=a[0],column=9).value=heure_arrivee
    wb.save()
