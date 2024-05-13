import openpyxl as xl
import sqlite3

def ecriture_excel(corps, id_aeronef):              
        ### Fonction qui inscrit le mail dans le fichier Excel ###
        #Ouverture du fichier
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']
        #Ligne excel
        i=6
        while feuille.cell(i, 4).value != None :
            i+=1

        # Identifiant aerodrome de depart
        ligne=corps[4].split('-')
        depart=ligne[1]

        #Recuperation vol dans bdd
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''SELECT "Heure de depart","Duree du vol", "Aerodrome d'arrivee", "Heure d'arrivee", "Chemin" FROM "Plans de vols" WHERE Aeronef = ? ''', (id_aeronef,))
        
        vol=[]
        for ligne in cur.fetchall():
            vol=list(ligne)
        
        cur.close()
        conn.close()

        #Ecriture dans le fichier excel
        feuille.cell(i,4).value = id_aeronef
        feuille.cell(i,5).value = depart[0:5]
        feuille.cell(i,6).value = vol[0]
        feuille.cell(i,7).value = vol[1]
        feuille.cell(i,8).value = vol[2]
        feuille.cell(i,9).value = vol[3]
        feuille.cell(i,10).value = vol[4]

        #Sauvegarder
        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')

def message_delai(corps,id_aeronef):               

        #Base de donnee
        ligne=corps[4].split('-')
        depart=ligne[1]
        
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()
        cur.execute('''UPDATE "Plans de vols" SET "Heure de depart" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))
        
        ligne2=corps[8].split('-')
        arrivee=ligne2[1]

        heure=int(depart[6:8])+int(arrivee[6:8])
        minute=int(depart[8:10])+int(arrivee[8:10])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)
        
        cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ? AND "Aerodrome de depart" = ?''', (depart[5:10],id_aeronef,depart[0:5]))

        conn.commit()
        conn.close()

        #Excel   
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']

        
        for row in feuille.iter_rows():
             for cell in row:
                 if cell.value == id_aeronef:
                     a = (cell.row,cell.column)
        
        feuille.cell(row=a[0],column=6).value=depart[5:10]

        feuille.cell(row=a[0],column=9).value=heure_arrivee

        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')

def message_changement(corps,id_aeronef):             

    conn = sqlite3.connect('/Users/thibautdejean/Desktop/vols_pai.db')
    cur = conn.cursor()

    cur.execute('''SELECT "Duree du vol" FROM "Plans de vols WHERE Aeronef = ? ''', (id_aeronef,))
    duree=cur.fetchall()

    ligne=corps[0].split('-')
    depart = ligne[2]
    arrivee = ligne[3]
    

    heure=int(depart[6:8])+int(duree[0:2])
    minute=int(depart[8:10])+int(duree[2:4])

    if int(minute)>60:
        minute=int(minute)-60
        heure+=1
    heure_arrivee = str(heure)+str(minute)
    
    
    cur.execute('''UPDATE "Plans de vols" SET "Aerodrome de depart" = ? WHERE Aeronef = ? ''', (depart[5:10],id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Heure de départ" = ? WHERE Aeronef = ? ''', (depart[0:5],id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Aerodrome d'arrivee" = ? WHERE Aeronef = ? ''', (arrivee,id_aeronef))
    cur.execute('''UPDATE "Plans de vols" SET "Heure d'arrivee" = ? WHERE Aeronef = ? ''', (heure_arrivee,id_aeronef))
    conn.commit()
    conn.close()

    #Excel   
    wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
    feuille = wb['Vols en cours']


    for row in feuille.iter_rows():
            for cell in row:
                if cell.value == id_aeronef:
                    a = (cell.row,cell.column)

    feuille.cell(row=a[0],column=5).value=depart[0:5]
    feuille.cell(row=a[0],column=6).value=depart[5:10]

    feuille.cell(row=a[0],column=8).value=arrivee
    feuille.cell(row=a[0],column=9).value=heure_arrivee
    wb.save()
       
def message_annulation(corps,id_aeronef):          
        #Base de donnee
        conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
        cur = conn.cursor()

        cur.execute('''DELETE FROM "Plans de vols" WHERE Aeronef = ?''', (id_aeronef,))

        conn.commit()
        conn.close()

        #Fichier excel
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']

        
        for row in feuille.iter_rows():
             for cell in row:
                 if cell.value == id_aeronef:
                     a = (cell.row,cell.column)
        
        for j in range(4,11):
            feuille.cell(row = a[0], column = j).value = None
            fill = xl.PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
            feuille.cell(row = a[0], column = j).fill = fill

        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
                       
def message_depart(corps,id_aeronef):               

        conn = sqlite3.connect('/Users/thibautdejean/Desktop/vols_pai.db')
        cur = conn.cursor()
               
        # Identification de l'aeronef
        ligne=corps[0].split('-')
        b = ligne[1].split(' ')
        id = ' '+b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]+' '
        print(id)

        cur.execute('''SELECT "Duree du vol" FROM "Plans de vols WHERE Aeronef = ? ''', (id,))
        duree=cur.fetchall()


        #Heure de départ et d'arrivée : 
        ligne=corps[0].split('-')
        depart = ligne[2]
    

        heure=int(depart[6:8])+int(duree[0:2])
        minute=int(depart[8:10])+int(duree[2:4])

        if int(minute)>60:
            minute=int(minute)-60
            heure+=1
        heure_arrivee = str(heure)+str(minute)
    
        # Changement de couleur sur l'excel
        wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
        feuille = wb['Vols en cours']

        
        a=[1,1]
        for row in feuille.iter_rows():
             for cell in row :
                 if str(cell.value) == str(id) : 
                    a = (cell.column,cell.row)
                    print(a)

        
        feuille.cell(row=a[0],column=6).value=depart[5:10]
        feuille.cell(row=a[0],column=9).value=heure_arrivee
         
        for j in range(4,11):
            fill = xl.styles.PatternFill(start_color="FF00FF00", end_color="FF00FF00", patternType='solid')            
            feuille.cell(row = a[0], column = j).fill = fill

        wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')

def message_arrivee(corps,id_aeronef):              
        
    # Identification de l'aeronef
    ligne=corps[0].split('-')
    b = ligne[1].split(' ')
    id = ' '+b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]+' '
    idbdd = b[0][0:2]+b[0][len(b[0])-2:len(b[0])]+b[1]
    

    # Supression BDD

    conn = sqlite3.connect('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols_pai_3.db')
    cur = conn.cursor()

    cur.execute('''DELETE FROM "Plans de vols" WHERE Aeronef = ? ''', (idbdd,))

    conn.commit()
    conn.close()

    # Suppression ligne Excel

    wb = xl.load_workbook('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
    feuille = wb['Vols en cours']

    for row in feuille.iter_rows():
            for cell in row :
                if str(cell.value) == str(id) : 
                    ligne = cell.row
                
    for j in range(4,11):
        feuille.cell(ligne,j).value = None
        feuille.cell(ligne,j).fill = xl.styles.PatternFill(fill_type=None)


    wb.save('/Users/thibautdejean/Downloads/PAI-git/PAI-3/vols.xlsx')
